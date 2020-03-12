import rpa as r
from openpyxl import load_workbook
r.init() #Inicialization
r.url('http://www.rpachallenge.com') # Access URL
r.download('http://www.rpachallenge.com/assets/downloadFiles/challenge.xlsx', 'challenge.xlsx') # Download File
wb = load_workbook(filename = 'challenge.xlsx', read_only=True) # Open file
ws = wb.active
r.type('Start', '[enter]')
lista = ["FirstName", "LastName", "CompanyName", "Role", "Address", "Email", "Phone"]
for i in range (2,12):
    cont = 0
    for j in range(1,8):
        cell_obj=ws.cell(row=i,column=j)
        r.type("//*[@ng-reflect-name=\"label{}\"]".format(lista[cont]), str(cell_obj.value))
        cont += 1   
    r.type('submit', '[enter]')
